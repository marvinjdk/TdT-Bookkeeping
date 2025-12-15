from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, status
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Literal
import uuid
from datetime import datetime, timezone, timedelta
from passlib.context import CryptContext
from jose import JWTError, jwt
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Security
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
SECRET_KEY = os.environ.get('SECRET_KEY', 'your-secret-key-change-in-production')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_DAYS = 30
security = HTTPBearer()

# Create the main app
app = FastAPI()
api_router = APIRouter(prefix="/api")

# Models
class UserCreate(BaseModel):
    username: str
    password: str
    role: Literal["admin", "afdeling", "superbruger"]
    afdeling_navn: Optional[str] = None

class UserLogin(BaseModel):
    username: str
    password: str

class UserPasswordUpdate(BaseModel):
    new_password: str

class UserAfdelingUpdate(BaseModel):
    afdeling_navn: str

class AfdelingCreate(BaseModel):
    navn: str

class Afdeling(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    navn: str
    oprettet: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    role: str
    afdeling_navn: Optional[str] = None

class Token(BaseModel):
    access_token: str
    token_type: str
    user: User

class SettingsUpdate(BaseModel):
    startsaldo: float = 0.0
    periode_start: str = "01-10-2024"
    periode_slut: str = "30-09-2025"
    regnskabsaar: str = "2024-2025"

class SettingsModel(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    afdeling_id: str
    startsaldo: float = 0.0
    periode_start: str = "01-10-2024"
    periode_slut: str = "30-09-2025"
    regnskabsaar: str = "2024-2025"
    naeste_bilagnr: int = 1

class TransactionCreate(BaseModel):
    bank_dato: str
    tekst: str
    formal: str
    belob: float
    type: Literal["indtaegt", "udgift"]

class Transaction(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    afdeling_id: str
    bilagnr: str
    bank_dato: str
    tekst: str
    formal: str
    belob: float
    type: str
    regnskabsaar: Optional[str] = None
    kvittering_url: Optional[str] = None
    oprettet: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class AfdelingSaldo(BaseModel):
    afdeling_id: str
    afdeling_navn: str
    aktuelt_saldo: float

class DashboardStats(BaseModel):
    model_config = ConfigDict(extra="ignore")
    aktuelt_saldo: Optional[float] = None
    total_indtaegter: float
    total_udgifter: float
    antal_posteringer: Optional[int] = None
    afdelinger_saldi: Optional[List[AfdelingSaldo]] = None

# Helper functions
def hash_password(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(days=ACCESS_TOKEN_EXPIRE_DAYS)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Ugyldig token")
        
        user = await db.users.find_one({"id": user_id}, {"_id": 0})
        if user is None:
            raise HTTPException(status_code=401, detail="Bruger ikke fundet")
        return User(**user)
    except JWTError:
        raise HTTPException(status_code=401, detail="Ugyldig token")

# Auth routes
@api_router.post("/auth/login", response_model=Token)
async def login(credentials: UserLogin):
    user = await db.users.find_one({"username": credentials.username}, {"_id": 0})
    if not user or not verify_password(credentials.password, user["password"]):
        raise HTTPException(status_code=401, detail="Ugyldigt brugernavn eller adgangskode")
    
    user_obj = User(**user)
    access_token = create_access_token(data={"sub": user_obj.id})
    return Token(access_token=access_token, token_type="bearer", user=user_obj)

@api_router.get("/auth/me", response_model=User)
async def get_me(current_user: User = Depends(get_current_user)):
    return current_user

# Admin routes
@api_router.post("/admin/users", response_model=User)
async def create_user(user_data: UserCreate, current_user: User = Depends(get_current_user)):
    if current_user.role != "superbruger":
        raise HTTPException(status_code=403, detail="Kun superbruger kan oprette brugere")
    
    existing = await db.users.find_one({"username": user_data.username}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Brugernavn eksisterer allerede")
    
    user_dict = user_data.model_dump()
    user_dict["password"] = hash_password(user_dict["password"])
    user_obj = User(**{k: v for k, v in user_dict.items() if k != "password"})
    
    doc = user_obj.model_dump()
    doc["password"] = user_dict["password"]
    await db.users.insert_one(doc)
    return user_obj

@api_router.get("/admin/users", response_model=List[User])
async def list_users(current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "superbruger"]:
        raise HTTPException(status_code=403, detail="Kun admins kan se alle brugere")
    
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    return [User(**u) for u in users]

@api_router.delete("/admin/users/{user_id}")
async def delete_user(user_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role != "superbruger":
        raise HTTPException(status_code=403, detail="Kun superbruger kan slette brugere")
    
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Bruger ikke fundet")
    return {"success": True}

@api_router.put("/admin/users/{user_id}/password")
async def update_user_password(
    user_id: str, 
    password_update: UserPasswordUpdate,
    current_user: User = Depends(get_current_user)
):
    if current_user.role != "superbruger":
        raise HTTPException(status_code=403, detail="Kun superbruger kan ændre passwords")
    
    hashed_password = hash_password(password_update.new_password)
    result = await db.users.update_one(
        {"id": user_id},
        {"$set": {"password": hashed_password}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Bruger ikke fundet")
    return {"success": True}

@api_router.put("/admin/users/{user_id}/afdeling")
async def update_user_afdeling(
    user_id: str, 
    afdeling_update: UserAfdelingUpdate,
    current_user: User = Depends(get_current_user)
):
    if current_user.role != "superbruger":
        raise HTTPException(status_code=403, detail="Kun superbruger kan ændre afdelingsnavn")
    
    # Check if user is an afdeling
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user or user.get("role") != "afdeling":
        raise HTTPException(status_code=400, detail="Kun afdelinger kan få ændret navn")
    
    result = await db.users.update_one(
        {"id": user_id},
        {"$set": {"afdeling_navn": afdeling_update.afdeling_navn}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Bruger ikke fundet")
    return {"success": True}

# Afdelinger endpoints
@api_router.get("/admin/afdelinger", response_model=List[Afdeling])
async def list_afdelinger(current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "superbruger"]:
        raise HTTPException(status_code=403, detail="Kun admins kan se afdelinger")
    
    afdelinger = await db.afdelinger.find({}, {"_id": 0}).sort("navn", 1).to_list(1000)
    return [Afdeling(**a) for a in afdelinger]

@api_router.post("/admin/afdelinger", response_model=Afdeling)
async def create_afdeling(afdeling: AfdelingCreate, current_user: User = Depends(get_current_user)):
    if current_user.role != "superbruger":
        raise HTTPException(status_code=403, detail="Kun superbruger kan oprette afdelinger")
    
    # Check if afdeling already exists
    existing = await db.afdelinger.find_one({"navn": afdeling.navn}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Afdeling findes allerede")
    
    afdeling_obj = Afdeling(navn=afdeling.navn)
    await db.afdelinger.insert_one(afdeling_obj.model_dump())
    return afdeling_obj

@api_router.delete("/admin/afdelinger/{afdeling_id}")
async def delete_afdeling(afdeling_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role != "superbruger":
        raise HTTPException(status_code=403, detail="Kun superbruger kan slette afdelinger")
    
    # Check if any users are using this afdeling
    users_with_afdeling = await db.users.find_one({"afdeling_navn": {"$exists": True}}, {"_id": 0})
    afdeling_to_delete = await db.afdelinger.find_one({"id": afdeling_id}, {"_id": 0})
    
    if afdeling_to_delete and users_with_afdeling:
        # Check if this specific afdeling name is in use
        in_use = await db.users.find_one({"afdeling_navn": afdeling_to_delete["navn"]}, {"_id": 0})
        if in_use:
            raise HTTPException(status_code=400, detail="Kan ikke slette afdeling der er i brug")
    
    result = await db.afdelinger.delete_one({"id": afdeling_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Afdeling ikke fundet")
    return {"success": True}


# Historiske data endpoints
@api_router.get("/historik/regnskabsaar")
async def get_available_regnskabsaar(current_user: User = Depends(get_current_user)):
    """Get list of available regnskabsår for filtering historical data (admin only)"""
    if current_user.role not in ["admin", "superbruger"]:
        raise HTTPException(status_code=403, detail="Kun admins kan se historiske data")
    
    # Get unique regnskabsår from settings
    pipeline = [
        {"$group": {"_id": "$regnskabsaar"}},
        {"$sort": {"_id": -1}}
    ]
    
    results = await db.settings.aggregate(pipeline).to_list(None)
    regnskabsaar_list = [r["_id"] for r in results if r["_id"]]
    
    return {"regnskabsaar": regnskabsaar_list}

@api_router.get("/admin/settings/all")
async def get_all_settings(current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "superbruger"]:
        raise HTTPException(status_code=403, detail="Kun admins kan se alle indstillinger")
    
    # Get all afdelinger
    afdelinger = await db.users.find({"role": "afdeling"}, {"_id": 0}).to_list(100)
    
    result = []
    for afdeling in afdelinger:
        settings = await db.settings.find_one({"afdeling_id": afdeling["id"]}, {"_id": 0})
        if not settings:
            settings_obj = SettingsModel(afdeling_id=afdeling["id"])
            await db.settings.insert_one(settings_obj.model_dump())
            settings = settings_obj.model_dump()
        
        result.append({
            "afdeling_id": afdeling["id"],
            "afdeling_navn": afdeling["afdeling_navn"],
            "settings": SettingsModel(**settings)
        })
    
    return result

@api_router.put("/admin/settings/{afdeling_id}")
async def update_afdeling_settings(
    afdeling_id: str,
    settings_update: SettingsUpdate,
    current_user: User = Depends(get_current_user)
):
    if current_user.role not in ["admin", "superbruger"]:
        raise HTTPException(status_code=403, detail="Kun admins kan opdatere indstillinger")
    
    # Get existing settings or create new
    existing = await db.settings.find_one({"afdeling_id": afdeling_id}, {"_id": 0})
    if existing:
        settings_obj = SettingsModel(**existing)
    else:
        settings_obj = SettingsModel(afdeling_id=afdeling_id)
    
    # Update with new values
    update_data = settings_update.model_dump()
    for key, value in update_data.items():
        setattr(settings_obj, key, value)
    
    await db.settings.update_one(
        {"afdeling_id": afdeling_id},
        {"$set": settings_obj.model_dump()},
        upsert=True
    )
    return settings_obj

# Settings routes
@api_router.get("/settings", response_model=SettingsModel)
async def get_settings(current_user: User = Depends(get_current_user)):
    afdeling_id = current_user.id if current_user.role == "afdeling" else None
    if not afdeling_id:
        raise HTTPException(status_code=400, detail="Kun afdelinger har indstillinger")
    
    settings = await db.settings.find_one({"afdeling_id": afdeling_id}, {"_id": 0})
    if not settings:
        # Create default settings
        settings_obj = SettingsModel(afdeling_id=afdeling_id)
        await db.settings.insert_one(settings_obj.model_dump())
        return settings_obj
    return SettingsModel(**settings)

@api_router.put("/settings", response_model=SettingsModel)
async def update_settings(settings_update: SettingsUpdate, current_user: User = Depends(get_current_user)):
    if current_user.role != "afdeling":
        raise HTTPException(status_code=403, detail="Kun afdelinger kan opdatere indstillinger")
    
    # Get existing settings or create new
    existing = await db.settings.find_one({"afdeling_id": current_user.id}, {"_id": 0})
    if existing:
        settings_obj = SettingsModel(**existing)
    else:
        settings_obj = SettingsModel(afdeling_id=current_user.id)
    
    # Update with new values
    update_data = settings_update.model_dump()
    for key, value in update_data.items():
        setattr(settings_obj, key, value)
    
    await db.settings.update_one(
        {"afdeling_id": current_user.id},
        {"$set": settings_obj.model_dump()},
        upsert=True
    )
    return settings_obj

# Transaction routes
@api_router.post("/transactions", response_model=Transaction)
async def create_transaction(transaction: TransactionCreate, current_user: User = Depends(get_current_user)):
    if current_user.role != "afdeling":
        raise HTTPException(status_code=403, detail="Kun afdelinger kan oprette posteringer")
    
    # Get or create settings to get next bilagnr
    settings = await db.settings.find_one({"afdeling_id": current_user.id}, {"_id": 0})
    if not settings:
        settings_obj = SettingsModel(afdeling_id=current_user.id)
        await db.settings.insert_one(settings_obj.model_dump())
        naeste_bilagnr = 1
    else:
        naeste_bilagnr = settings.get("naeste_bilagnr", 1)
    
    # Generate bilagnr
    bilagnr = f"B{str(naeste_bilagnr).zfill(3)}"
    
    # Update next bilagnr in settings
    await db.settings.update_one(
        {"afdeling_id": current_user.id},
        {"$set": {"naeste_bilagnr": naeste_bilagnr + 1}},
        upsert=True
    )
    
    trans_dict = transaction.model_dump()
    trans_dict["bilagnr"] = bilagnr
    
    # Automatically assign regnskabsaar from settings
    if settings:
        trans_dict["regnskabsaar"] = settings.get("regnskabsaar", "2024-2025")
    else:
        trans_dict["regnskabsaar"] = "2024-2025"
    
    trans_obj = Transaction(afdeling_id=current_user.id, **trans_dict)
    await db.transactions.insert_one(trans_obj.model_dump())
    return trans_obj

@api_router.get("/transactions", response_model=List[Transaction])
async def list_transactions(
    afdeling_id: Optional[str] = None,
    regnskabsaar: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    query = {}
    if current_user.role == "afdeling":
        query["afdeling_id"] = current_user.id
    elif current_user.role == "admin" and afdeling_id:
        query["afdeling_id"] = afdeling_id
    
    # Filter by regnskabsaar if provided (admin only)
    if regnskabsaar and current_user.role in ["admin", "superbruger"]:
        query["regnskabsaar"] = regnskabsaar
    
    projection = {
        "_id": 0, "id": 1, "afdeling_id": 1, "bilagnr": 1, 
        "bank_dato": 1, "tekst": 1, "formal": 1, "belob": 1, 
        "type": 1, "regnskabsaar": 1, "kvittering_url": 1, "oprettet": 1
    }
    transactions = await db.transactions.find(query, projection).sort("bank_dato", -1).to_list(1000)
    return [Transaction(**t) for t in transactions]

@api_router.get("/transactions/{transaction_id}", response_model=Transaction)
async def get_transaction(transaction_id: str, current_user: User = Depends(get_current_user)):
    transaction = await db.transactions.find_one({"id": transaction_id}, {"_id": 0})
    if not transaction:
        raise HTTPException(status_code=404, detail="Postering ikke fundet")
    
    if current_user.role == "afdeling" and transaction["afdeling_id"] != current_user.id:
        raise HTTPException(status_code=403, detail="Ingen adgang")
    
    return Transaction(**transaction)

@api_router.put("/transactions/{transaction_id}", response_model=Transaction)
async def update_transaction(
    transaction_id: str,
    transaction: TransactionCreate,
    current_user: User = Depends(get_current_user)
):
    existing = await db.transactions.find_one({"id": transaction_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Postering ikke fundet")
    
    if current_user.role == "afdeling" and existing["afdeling_id"] != current_user.id:
        raise HTTPException(status_code=403, detail="Ingen adgang")
    
    # Don't allow bilagnr to be updated - keep existing
    update_data = transaction.model_dump()
    # Bilagnr is already set and should not change
    await db.transactions.update_one({"id": transaction_id}, {"$set": update_data})
    
    updated = await db.transactions.find_one({"id": transaction_id}, {"_id": 0})
    return Transaction(**updated)

@api_router.delete("/transactions/{transaction_id}")
async def delete_transaction(transaction_id: str, current_user: User = Depends(get_current_user)):
    existing = await db.transactions.find_one({"id": transaction_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Postering ikke fundet")
    
    if current_user.role == "afdeling" and existing["afdeling_id"] != current_user.id:
        raise HTTPException(status_code=403, detail="Ingen adgang")
    
    result = await db.transactions.delete_one({"id": transaction_id})
    return {"success": True}

@api_router.post("/transactions/{transaction_id}/upload")
async def upload_receipt(
    transaction_id: str,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    transaction = await db.transactions.find_one({"id": transaction_id}, {"_id": 0})
    if not transaction:
        raise HTTPException(status_code=404, detail="Postering ikke fundet")
    
    if current_user.role == "afdeling" and transaction["afdeling_id"] != current_user.id:
        raise HTTPException(status_code=403, detail="Ingen adgang")
    
    # Save file locally (later can be Dropbox)
    upload_dir = Path("/app/uploads")
    upload_dir.mkdir(exist_ok=True)
    
    file_extension = Path(file.filename).suffix
    filename = f"{transaction_id}_{uuid.uuid4()}{file_extension}"
    file_path = upload_dir / filename
    
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)
    
    kvittering_url = f"/uploads/{filename}"
    await db.transactions.update_one(
        {"id": transaction_id},
        {"$set": {"kvittering_url": kvittering_url}}
    )
    
    return {"success": True, "url": kvittering_url}

# Dashboard stats
@api_router.get("/dashboard/stats", response_model=DashboardStats)
async def get_dashboard_stats(
    afdeling_id: Optional[str] = None,
    regnskabsaar: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    # Admin sees all afdelinger with their saldi
    if current_user.role in ["admin", "superbruger"] and not afdeling_id:
        # Get all afdelinger from the afdelinger collection (not just users)
        afdelinger = await db.afdelinger.find({}, {"_id": 0}).to_list(100)
        
        afdelinger_saldi = []
        total_indtaegter_all = 0.0
        total_udgifter_all = 0.0
        total_startsaldo_all = 0.0
        
        for afdeling in afdelinger:
            # Get transactions for this afdeling (match by navn)
            # First, find any user with this afdeling_navn to get the afdeling_id used in transactions
            user_with_afdeling = await db.users.find_one(
                {"afdeling_navn": afdeling["navn"], "role": "afdeling"}, 
                {"_id": 0, "id": 1}
            )
            
            # Use the afdeling_id from the user, or the afdeling id itself
            afdeling_id_for_query = user_with_afdeling["id"] if user_with_afdeling else afdeling["id"]
            
            # Build match query with optional regnskabsaar filter
            match_query = {"afdeling_id": afdeling_id_for_query}
            if regnskabsaar:
                match_query["regnskabsaar"] = regnskabsaar
            
            pipeline = [
                {"$match": match_query},
                {
                    "$group": {
                        "_id": "$type",
                        "total": {"$sum": "$belob"}
                    }
                }
            ]
            
            results = await db.transactions.aggregate(pipeline).to_list(None)
            
            indtaegter = 0.0
            udgifter = 0.0
            
            for result in results:
                if result["_id"] == "indtaegt":
                    indtaegter = result["total"]
                elif result["_id"] == "udgift":
                    udgifter = result["total"]
            
            # Get startsaldo
            settings = await db.settings.find_one({"afdeling_id": afdeling_id_for_query}, {"_id": 0})
            startsaldo = settings.get("startsaldo", 0.0) if settings else 0.0
            
            aktuelt_saldo = startsaldo + indtaegter - udgifter
            
            afdelinger_saldi.append(AfdelingSaldo(
                afdeling_id=afdeling["id"],
                afdeling_navn=afdeling["navn"],
                aktuelt_saldo=aktuelt_saldo
            ))
            
            total_indtaegter_all += indtaegter
            total_udgifter_all += udgifter
            total_startsaldo_all += startsaldo
        
        # Calculate total current balance
        total_aktuelt_saldo = total_startsaldo_all + total_indtaegter_all - total_udgifter_all
        
        return DashboardStats(
            aktuelt_saldo=total_aktuelt_saldo,
            total_indtaegter=total_indtaegter_all,
            total_udgifter=total_udgifter_all,
            afdelinger_saldi=afdelinger_saldi
        )
    else:
        # Single afdeling view
        query = {}
        target_afdeling_id = None
        
        if current_user.role == "afdeling":
            target_afdeling_id = current_user.id
            query["afdeling_id"] = current_user.id
        elif current_user.role in ["admin", "superbruger"] and afdeling_id:
            target_afdeling_id = afdeling_id
            query["afdeling_id"] = afdeling_id
        
        # Add regnskabsaar filter if provided
        if regnskabsaar:
            query["regnskabsaar"] = regnskabsaar
        
        # Use aggregation pipeline for efficient calculation
        pipeline = [
            {"$match": query},
            {
                "$group": {
                    "_id": "$type",
                    "total": {"$sum": "$belob"},
                    "count": {"$sum": 1}
                }
            }
        ]
        
        results = await db.transactions.aggregate(pipeline).to_list(None)
        
        total_indtaegter = 0.0
        total_udgifter = 0.0
        antal_posteringer = 0
        
        for result in results:
            if result["_id"] == "indtaegt":
                total_indtaegter = result["total"]
                antal_posteringer += result["count"]
            elif result["_id"] == "udgift":
                total_udgifter = result["total"]
                antal_posteringer += result["count"]
        
        # Get startsaldo
        startsaldo = 0.0
        if target_afdeling_id:
            settings = await db.settings.find_one({"afdeling_id": target_afdeling_id}, {"_id": 0})
            if settings:
                startsaldo = settings.get("startsaldo", 0.0)
        
        aktuelt_saldo = startsaldo + total_indtaegter - total_udgifter
        
        return DashboardStats(
            aktuelt_saldo=aktuelt_saldo,
            total_indtaegter=total_indtaegter,
            total_udgifter=total_udgifter,
            antal_posteringer=antal_posteringer
        )

# Excel export
@api_router.get("/export/excel")
async def export_excel(
    afdeling_id: Optional[str] = None,
    regnskabsaar: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    wb = Workbook()
    
    # If admin and no specific afdeling_id, export all with separate sheets
    if current_user.role == "admin" and not afdeling_id:
        # Get all afdelinger
        afdelinger = await db.users.find({"role": "afdeling"}, {"_id": 0}).to_list(100)
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheet for each afdeling
        all_transactions = []
        for afdeling in afdelinger:
            await create_afdeling_sheet(wb, afdeling["id"], afdeling["afdeling_navn"])
            
            # Collect for combined sheet
            projection = {
                "_id": 0, "bilagnr": 1, "bank_dato": 1, 
                "tekst": 1, "formal": 1, "belob": 1, "type": 1, "afdeling_id": 1
            }
            trans = await db.transactions.find(
                {"afdeling_id": afdeling["id"]}, projection
            ).sort("bank_dato", 1).to_list(10000)
            
            for t in trans:
                t["afdeling_navn"] = afdeling["afdeling_navn"]
                all_transactions.append(t)
        
        # Create combined sheet
        ws_combined = wb.create_sheet("Alle hold")
        ws_combined.append(["Hold", "Bilagnr.", "Bank dato", "Tekst", "Formål", "Beløb", "Type"])
        
        # Style headers
        header_fill = PatternFill(start_color="109848", end_color="109848", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws_combined[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Sort all by date and add data
        all_transactions.sort(key=lambda x: x["bank_dato"])
        for t in all_transactions:
            ws_combined.append([
                t["afdeling_navn"],
                t["bilagnr"],
                t["bank_dato"],
                t["tekst"],
                t["formal"],
                t["belob"],
                t["type"]
            ])
        
        auto_adjust_columns(ws_combined)
    else:
        # Single afdeling export
        target_afdeling_id = current_user.id if current_user.role == "afdeling" else afdeling_id
        
        # Get afdeling name
        afdeling_user = await db.users.find_one({"id": target_afdeling_id}, {"_id": 0})
        afdeling_navn = afdeling_user.get("afdeling_navn", "Bogføring") if afdeling_user else "Bogføring"
        
        wb.remove(wb.active)
        await create_afdeling_sheet(wb, target_afdeling_id, afdeling_navn)
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=tour_de_taxa_bogforing.xlsx"}
    )

async def create_afdeling_sheet(wb, afdeling_id, afdeling_navn, regnskabsaar=None):
    """Create a sheet for a specific afdeling with startsaldo and aktuel saldo"""
    # Clean sheet name - remove invalid characters and limit to 31 chars
    clean_name = afdeling_navn.replace("/", "-").replace("\\", "-").replace("[", "").replace("]", "")
    clean_name = clean_name.replace("*", "").replace("?", "").replace(":", "")[:31]
    ws = wb.create_sheet(clean_name)
    
    # Get settings for startsaldo
    settings = await db.settings.find_one({"afdeling_id": afdeling_id}, {"_id": 0})
    startsaldo = settings.get("startsaldo", 0.0) if settings else 0.0
    
    # Build query with optional regnskabsaar filter
    query = {"afdeling_id": afdeling_id}
    if regnskabsaar:
        query["regnskabsaar"] = regnskabsaar
    
    # Get transactions
    projection = {
        "_id": 0, "bilagnr": 1, "bank_dato": 1, 
        "tekst": 1, "formal": 1, "belob": 1, "type": 1
    }
    transactions = await db.transactions.find(query, projection).sort("bank_dato", 1).to_list(10000)
    
    # Calculate sums
    total_indtaegter = sum(t["belob"] for t in transactions if t["type"] == "indtaegt")
    total_udgifter = sum(t["belob"] for t in transactions if t["type"] == "udgift")
    aktuel_saldo = startsaldo + total_indtaegter - total_udgifter
    
    # Add startsaldo row
    ws.append(["Startsaldo", "", "", "", startsaldo, ""])
    ws[1][0].font = Font(bold=True)
    ws[1][4].font = Font(bold=True)
    
    # Add empty row
    ws.append([])
    
    # Headers
    headers = ["Bilagnr.", "Bank dato", "Tekst", "Formål", "Beløb", "Type"]
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="109848", end_color="109848", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_row = ws[3]
    for cell in header_row:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    for t in transactions:
        ws.append([
            t["bilagnr"],
            t["bank_dato"],
            t["tekst"],
            t["formal"],
            t["belob"],
            t["type"]
        ])
    
    # Add empty row
    ws.append([])
    
    # Add aktuel saldo row
    last_row = ws.max_row + 1
    ws.append(["Aktuel saldo", "", "", "", aktuel_saldo, ""])
    ws[last_row][0].font = Font(bold=True, color="109848")
    ws[last_row][4].font = Font(bold=True, color="109848")
    
    auto_adjust_columns(ws)

def auto_adjust_columns(ws):
    """Auto-adjust column widths"""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()